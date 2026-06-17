"""
cli/excel_writer.py — Incremental Excel writer using openpyxl.
Supports branch sections with title rows, column headers, data rows,
a subject code→name footer, and auto-saving.
"""

import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Style constants
FONT_TITLE = Font(name="Segoe UI", size=12, bold=True, color="1F497D")
FONT_HEADER = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
FONT_DATA = Font(name="Segoe UI", size=10)
FONT_FOOTER_LABEL = Font(name="Segoe UI", size=8, bold=True, color="1F497D")
FONT_FOOTER_VALUE = Font(name="Segoe UI", size=8, italic=True, color="595959")

FILL_HEADER = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
FILL_TITLE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_ALT_ROW = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
FILL_FOOTER = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# SGPA grade colors for cell fills
SGPA_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
SGPA_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SGPA_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

GRADE_FONT_FAIL = Font(name="Segoe UI", size=10, bold=True, color="CC0000")
GRADE_FONT_EXCELLENT = Font(name="Segoe UI", size=10, bold=True, color="006100")


class ExcelWriter:
    """
    Manages writing batch results to an Excel workbook incrementally.
    Each branch/semester group gets its own section with title, headers,
    data rows, and a subject code→name footer mapping.
    """

    def __init__(self, session):
        """Initialize a new workbook."""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Results"
        self.session = session
        self.current_row = 1
        self.current_subject_codes = []  # ordered list of subject codes for current section
        self.subject_name_map = {}       # code → name mapping for current section
        self.section_data_start_row = 1  # where data rows started (for footer placement)
        self.row_count_in_section = 0
        self.section_students = []       # store (roll_no, name, sgpa) for toppers list

        # Generate filename
        safe_session = session.replace("(", "").replace(")", "").replace("-", "_").replace(" ", "_")
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = f"BPUT_{safe_session}_{timestamp}.xlsx"

        os.makedirs("exports", exist_ok=True)
        self.filepath = os.path.join("exports", self.filename)

    def start_branch_section(self, branch, semester, subject_codes, subject_name_map):
        """
        Write a section title row and column headers for a new branch/semester group.

        Args:
            branch: branch name string (e.g., "B.Tech.(COMPUTER SCIENCE & ENGINEERING)")
            semester: semester string (e.g., "5th")
            subject_codes: ordered list of subject code strings
            subject_name_map: dict mapping subject code → subject name
        """
        # Leave 2 blank rows between sections (except the first)
        if self.current_row > 1:
            self.current_row += 2

        self.current_semester = semester
        self.current_subject_codes = subject_codes
        self.subject_name_map = subject_name_map
        self.row_count_in_section = 0
        self.section_students = []

        # Title row
        title_text = f"BRANCH: {branch} | SEMESTER: {semester} | SESSION: {self.session}"
        total_cols = 3 + len(subject_codes)  # Roll No, Name, SGPA, + subjects
        self.ws.merge_cells(
            start_row=self.current_row, start_column=1,
            end_row=self.current_row, end_column=max(total_cols, 4)
        )
        cell = self.ws.cell(row=self.current_row, column=1, value=title_text)
        cell.font = FONT_TITLE
        cell.fill = FILL_TITLE
        cell.alignment = ALIGN_LEFT
        self.ws.row_dimensions[self.current_row].height = 28
        self.current_row += 1

        # Column headers
        headers = ["Roll No", "Name", "SGPA"] + subject_codes
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=self.current_row, column=col_idx, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
        self.ws.row_dimensions[self.current_row].height = 22
        self.current_row += 1
        self.section_data_start_row = self.current_row

    def add_new_subject_header(self, code, name):
        """
        Add a new subject column header dynamically if encountered later.
        """
        if code in self.current_subject_codes:
            return
            
        self.current_subject_codes.append(code)
        self.subject_name_map[code] = name
        
        header_row = self.section_data_start_row - 1
        col_idx = 3 + len(self.current_subject_codes)
        
        cell = self.ws.cell(row=header_row, column=col_idx, value=code)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        
        try:
            self.ws.unmerge_cells(start_row=header_row - 1, start_column=1, end_row=header_row - 1, end_column=col_idx - 1)
            self.ws.merge_cells(start_row=header_row - 1, start_column=1, end_row=header_row - 1, end_column=col_idx)
        except Exception:
            pass

    def add_student_row(self, roll_no, name, sgpa, grades_dict):
        """
        Append a single student's data row.

        Args:
            roll_no: registration number string
            name: student name string
            sgpa: SGPA string (e.g., "6.73")
            grades_dict: dict mapping subject_code → grade letter
        """
        self.row_count_in_section += 1
        is_alt = self.row_count_in_section % 2 == 0
        
        try:
            sgpa_val = float(sgpa)
        except (ValueError, TypeError):
            sgpa_val = sgpa
            
        self.section_students.append((roll_no, name, sgpa_val))

        # Roll No
        try:
            roll_val = int(roll_no)
        except (ValueError, TypeError):
            roll_val = roll_no
            
        cell = self.ws.cell(row=self.current_row, column=1, value=roll_val)
        cell.font = FONT_DATA
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = FILL_ALT_ROW

        # Name
        cell = self.ws.cell(row=self.current_row, column=2, value=name)
        cell.font = FONT_DATA
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = FILL_ALT_ROW

        # SGPA (colored)
        cell = self.ws.cell(row=self.current_row, column=3, value=sgpa_val)
        cell.font = FONT_DATA
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        try:
            val = float(sgpa)
            if val >= 8.0:
                cell.fill = SGPA_FILL_GREEN
            elif val >= 6.0:
                cell.fill = SGPA_FILL_YELLOW
            else:
                cell.fill = SGPA_FILL_RED
        except (ValueError, TypeError):
            if is_alt:
                cell.fill = FILL_ALT_ROW

        # Subject grades
        for col_idx, code in enumerate(self.current_subject_codes, start=4):
            grade = grades_dict.get(code, "—")
            cell = self.ws.cell(row=self.current_row, column=col_idx, value=grade)
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER

            # Color the grade
            if grade in ("F", "M"):
                cell.font = GRADE_FONT_FAIL
            elif grade in ("O", "E"):
                cell.font = GRADE_FONT_EXCELLENT
            else:
                cell.font = FONT_DATA

            if is_alt and grade not in ("F", "M"):
                cell.fill = FILL_ALT_ROW

        self.ws.row_dimensions[self.current_row].height = 18
        self.current_row += 1


    def write_subject_footer(self):
        """
        Write a footer section that maps each subject code to its full name,
        and displays a Top 6 Toppers list at the top right of the section.
        """
        if not self.current_subject_codes:
            return

        # Identify top students
        valid_students = [s for s in self.section_students if isinstance(s[2], (int, float))]
        top_students = sorted(valid_students, key=lambda x: x[2], reverse=True)[:6]

        # --- Bottom Footer (Subject Codes & Toppers side-by-side) ---
        self.current_row += 1  # blank row
        start_row = self.current_row

        # Write Subject Code Reference Header (Columns 1-4)
        self.ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
        cell = self.ws.cell(row=start_row, column=1, value="Subject Code Reference:")
        cell.font = FONT_FOOTER_LABEL
        cell.fill = FILL_FOOTER
        cell.alignment = ALIGN_LEFT

        curr_sub_row = start_row + 1
        for code in self.current_subject_codes:
            name = self.subject_name_map.get(code, "Unknown")
            cell_code = self.ws.cell(row=curr_sub_row, column=1, value=code)
            cell_code.font = FONT_FOOTER_LABEL
            cell_code.fill = FILL_FOOTER
            cell_code.alignment = ALIGN_LEFT
            cell_code.border = THIN_BORDER

            self.ws.merge_cells(start_row=curr_sub_row, start_column=2, end_row=curr_sub_row, end_column=4)
            cell_name = self.ws.cell(row=curr_sub_row, column=2, value=name)
            cell_name.font = FONT_FOOTER_VALUE
            cell_name.fill = FILL_FOOTER
            cell_name.alignment = ALIGN_LEFT
            cell_name.border = THIN_BORDER
            curr_sub_row += 1

        # Write Toppers List right beside the Subject Codes (Columns 6-8)
        curr_top_row = start_row
        if top_students:
            topper_col = 6
            topper_row = start_row

            self.ws.merge_cells(start_row=topper_row, start_column=topper_col, end_row=topper_row, end_column=topper_col+2)
            title = f"Branch Toppers {self.session} {getattr(self, 'current_semester', '')} (Top 6):"
            cell = self.ws.cell(row=topper_row, column=topper_col, value=title)
            cell.font = FONT_FOOTER_LABEL
            cell.fill = FILL_TITLE
            cell.alignment = ALIGN_CENTER
            
            for i, h in enumerate(["Rank", "Name", "SGPA"]):
                c = self.ws.cell(row=topper_row+1, column=topper_col+i, value=h)
                c.font = FONT_FOOTER_LABEL
                c.fill = FILL_TITLE
                c.alignment = ALIGN_CENTER
                c.border = THIN_BORDER

            curr_top_row = topper_row + 2
            for rank, (r_no, name, sgpa) in enumerate(top_students, 1):
                c_rank = self.ws.cell(row=curr_top_row, column=topper_col, value=f"#{rank}")
                c_rank.font = FONT_FOOTER_LABEL
                c_rank.alignment = ALIGN_CENTER
                c_rank.border = THIN_BORDER
                
                c_name = self.ws.cell(row=curr_top_row, column=topper_col+1, value=name)
                c_name.font = FONT_FOOTER_VALUE
                c_name.alignment = ALIGN_LEFT
                c_name.border = THIN_BORDER

                c_sgpa = self.ws.cell(row=curr_top_row, column=topper_col+2, value=sgpa)
                c_sgpa.font = FONT_FOOTER_LABEL
                c_sgpa.alignment = ALIGN_CENTER
                c_sgpa.border = THIN_BORDER
                if sgpa >= 8.0:
                    c_sgpa.fill = SGPA_FILL_GREEN
                elif sgpa >= 6.0:
                    c_sgpa.fill = SGPA_FILL_YELLOW
                else:
                    c_sgpa.fill = SGPA_FILL_RED
                
                curr_top_row += 1

        # Advance current row past whichever block was longer
        self.current_row = max(curr_sub_row, curr_top_row)

    def rename_sheet(self, name):
        """Rename the current active worksheet."""
        self.ws.title = name[:31]

    def new_sheet(self, name):
        """Create a new worksheet tab and switch to it."""
        self.ws = self.wb.create_sheet(title=name[:31])  # Excel limits sheet name to 31 chars
        self.current_row = 1
        self.current_subject_codes = []
        self.subject_name_map = {}
        self.row_count_in_section = 0
        self.section_students = []

    def auto_fit_columns(self):
        """Auto-adjust column widths based on content."""
        for ws in self.wb.worksheets:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value and not isinstance(cell.value, (int, float)):
                            if len(str(cell.value)) > max_len:
                                max_len = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max(min(max_len + 4, 40), 10)

    def save(self):
        """Save the workbook to disk."""
        try:
            self.auto_fit_columns()
            self.wb.save(self.filepath)
        except Exception:
            # Silently handle save errors during auto-save
            pass

    def close(self):
        """Final save and cleanup."""
        self.auto_fit_columns()
        self.wb.save(self.filepath)

    def create_new_file(self, session):
        """
        Close the current workbook and create a brand new file.
        Returns the new ExcelWriter instance.
        """
        self.close()
        return ExcelWriter(session)
